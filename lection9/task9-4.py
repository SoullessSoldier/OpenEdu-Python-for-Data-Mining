"""
https://data.mos.ru/opendata/893 - неверный датаст для этой задачи, надо брать отсюда:
https://data.mos.ru/opendata/1251

Вам дан xlsx-файл с информацией о спортивных площадках в Москве. Вам необходимо сдать на проверку json-файл,
в котором будет храниться один словарь, ключами которого будут административные округа (AdmArea),
а значениями словари, в которых, в свою очередь, ключами будут названия районов (District),
относящихся к этому административному округу, а значениями - списки адресов площадок (Address) в том порядке, в котором они встречались в исходном файле.

Ваш файл должен выглядеть примерно так:

{"Северо-Западный административный округ": {"район Строгино": ["улица Исаковского, дом 24, корпус 1", "Неманский проезд, дом 9"],
"район Северное Тушино": ["улица Свободы, дом 56", "улица Свободы, дом 56", "улица Свободы, дом 56", "улица Свободы, дом 56"],
"район Покровское-Стрешнево": ["Иваньковское шоссе, дом 6"], "район Щукино": ["Сосновая улица, дом 3, строение 2"]}, ...}
data-25342-2019-09-30.xlsx - файл для датасета 893, 
правильный фал для задачи и датасета 1251 - data-25290-2019-09-30.xlsx

"""
from openpyxl import load_workbook
import json
wb=load_workbook('data-25290-2019-09-30.xlsx')
admareas={}
districts={}
adresses=[]
sheet=wb['Sheet0']
"""
admarea - row i, column 5
district - row i, column 6
address - row i, column 7
"""
for i in range(2,sheet.max_row+1):
    admarea=sheet.cell(row=i,column=5).value
    district=sheet.cell(row=i,column=6).value
    address=sheet.cell(row=i,column=7).value
    id=sheet.cell(row=i,column=1).value

    if admarea not in admareas:
        admareas[admarea]={}
    if district not in admareas[admarea]:
        admareas[admarea][district]=[]
    admareas[admarea][district].append(address)
fout=open("task9-4-out.json","w",encoding="utf-8")
#json.dump(admareas,fout,ensure_ascii=False,indent=4)
json.dump(admareas,fout,ensure_ascii=False)
fout.close()
