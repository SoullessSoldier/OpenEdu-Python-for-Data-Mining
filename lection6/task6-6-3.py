a='''
Васю назначили завхозом в туристической группе и он подошёл к подготовке ответственно, 
составив справочник продуктов с указанием калорийности на 100 грамм, 
а также содержание белков, жиров и углеводов на 100 грамм продукта. 
Ему не удалось найти всю информацию, поэтому некоторые ячейки остались незаполненными (можно считать их значение равным нулю). 
Также он использовал какой-то странный офисный пакет и разделял целую и дробную часть чисел запятой. 
Таблица доступна по ссылке https://stepik.org/media/attachments/lesson/258931/trekking3.xlsx

Вася составил раскладку по продуктам на весь поход (она на листе "Раскладка") 
с указанием номера дня, названия продукта и его количества в граммах. 
Для каждого дня посчитайте 4 числа: суммарную калорийность и граммы белков, жиров и углеводов. 
Числа округлите до целых вниз и введите через пробел. 
Информация о каждом дне должна выводиться в отдельной строке.
'''

def replace_none(s):
    s=str(s)
    s.replace(',','.')
    if s == 'None':
        s=0
    s=float(s)
    return s

from math import floor

from openpyxl import load_workbook
wb=load_workbook('trekking6-6-3.xlsx')
products=dict()
sheet=wb['Справочник']
for i in range(2,sheet.max_row+1):
    products[sheet.cell(row=i,column=1).value]=[replace_none(sheet.cell(row=i,column=2).value),
    replace_none(sheet.cell(row=i,column=3).value),
    replace_none(sheet.cell(row=i,column=4).value),
    replace_none(sheet.cell(row=i,column=5).value)]

sheet=wb['Раскладка']
raskladka=[]
for i in range(2,sheet.max_row+1):
    raskladka.append([sheet.cell(row=i,column=1).value,sheet.cell(row=i,column=2).value,sheet.cell(row=i,column=3).value/100])
#Нам надо найти номер минимального и максиммального дня, чтобы потом сделать range
min_day=sorted(raskladka)[0][0]
max_day=sorted(raskladka,reverse=True)[0][0]
print(min_day,max_day)
for day in range(min_day,max_day+1):
    food_list = []
    for raskl in raskladka:
        if raskl[0]==day:
            food_list.append([raskl[1], raskl[2] * products[raskl[1]][0], raskl[2] * products[raskl[1]][1],
                              raskl[2] * products[raskl[1]][2], raskl[2] * products[raskl[1]][3]])
    kcal = 0#килокалории
    b = 0#белки
    z = 0#жиры
    u = 0#углеводы
    for i in range(0, len(food_list)):
        kcal += food_list[i][1]
        b += food_list[i][2]
        z += food_list[i][3]
        u += food_list[i][4]
    print(*map(floor, (kcal, b, z, u)), sep=' ')

#Чертовы наркоманы в задании