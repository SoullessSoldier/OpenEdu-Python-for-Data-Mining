#Чтение XLSX-файлов
from openpyxl import load_workbook
wb=load_workbook('test2.xlsx')
#Узнать, какие лист есть в книге
#print(wb.get_sheet_names())#function deprecated
print(wb.sheetnames)
#sh=wb.get_sheet_by_name('Лист1')#depecated
sh=wb['Лист1']
print(sh.title)
print(sh['A1'].value)#тип value должен совпадать с типом содержимого ячейки
#навигация
c=sh['B2']
print(c.column)#столбец, нумерация с 1
print(c.row)#строка, нумерация с 1
print(c.coordinate)
#Либо так обратиться к ячейке
d=sh.cell(row=1,column=2)
print(d.value)