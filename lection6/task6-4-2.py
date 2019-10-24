a='''
Главный бухгалтер компании "Рога и копыта" случайно удалил ведомость с начисленной зарплатой. 
К счастью, у него сохранились расчётные листки всех сотрудников. 
Помогите по этим расчётным листкам восстановить зарплатную ведомость. 
Архив с расчётными листками доступен по ссылке https://stepik.org/media/attachments/lesson/245299/rogaikopyta.zip 
(вы можете скачать и распаковать его вручную или самостоятельно научиться делать это с помощью скрипта на Питоне).
Ведомость должна содержать 1000 строк, в каждой строке должно быть указано ФИО сотрудника и, через пробел, его зарплата. Сотрудники должны быть упорядочены по алфавиту.
'''
import openpyxl
import os

lst=[]
fout=open('out.txt','w',encoding='utf8')
for file in os.listdir("./rogaikopyta"):
    if file.endswith(".xlsx"):
        #print(file)
        wb = openpyxl.load_workbook(os.path.join("./rogaikopyta", file))
        sh = wb.active
        name = sh.cell(row=2, column=2).value
        zp = sh.cell(row=2, column=4).value
        lst.append([name,zp])
for i in sorted(lst):
    print(f'{i[0]} {i[1]}')
    fout.writelines(f'{i[0]} {i[1]}\n')
fout.close()