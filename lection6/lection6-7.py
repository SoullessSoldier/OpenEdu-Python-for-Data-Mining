#Запись в XLSX-файлы
from openpyxl import load_workbook
wb=load_workbook('test6-6-2.xlsx')
sheet=wb['Лист1']
#При записи файл не должен быт открыт/заблокирован другой программой!

sheet['A1']='Sasha'
sheet['A4']='Sasha'
#Пробуем добавить целый список
#Добавится после последней заполненной строки
sheet.append(['Dima',5])
wb.save('test6-6-2.xlsx')
#Задание - из лекции 6-6 выдать всем людям среднюю зарплату по региону
wb=load_workbook('test6-6-1.xlsx')
#sheet=wb.get_sheet_by_name('Лист1')
sheet=wb['Лист1']
salaries={}
where={}
#мы должны идти по всем строкам
for i in range(1,sheet.max_row+1):
    name=sheet.cell(row=i,column=1).value

    salary=sheet.cell(row=i,column=2).value
    salaries[name]=salary
#print(salaries)
#смотрим на регионы (Лист2)
sheet=wb['Лист2']
regions={}
for i in range(1,sheet.max_row+1):
    name=sheet.cell(row=i,column=1).value
    region=sheet.cell(row=i,column=2).value
    where[name]=region
    if region not in regions:
        regions[region]=[0,0]
    regions[region][0]+=1
    regions[region][1]+=salaries[name]

#regions[region][1]/regions[region][0]
sheet=wb['Лист1']
salary={}
for i in range(1,sheet.max_row+1):
    name = sheet.cell(row=i, column=1).value
    region=where[name]
    sheet.cell(row=i, column=2).value = regions[region][1]/regions[region][0]
wb.save('test6-6-3.xlsx')
#Рефакторинг!