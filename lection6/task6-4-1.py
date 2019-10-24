import openpyxl

wb = openpyxl.load_workbook('tab6-4-1.xlsx')
sh = wb.active
nmin = sh.cell(row=7, column=2).value
for rownum in range(8, 28):
    nmin = min(nmin, sh.cell(row=rownum, column=2).value)
print(nmin)