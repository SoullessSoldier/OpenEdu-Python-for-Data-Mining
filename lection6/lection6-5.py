#достать все содержимое таблицы
#если мы знаем размер таблицы
from openpyxl import load_workbook
wb=load_workbook('test2.xlsx')
sheet=wb['Лист1']
for i in range(1,4):
    for j in range(1,4):
        d=sheet.cell(row=i,column=j)
        print(d.value)

#если не знаем размер
#sheet.max_row
#sheet.max_column
for i in range(1,sheet.max_row+1):
    for j in range(1,sheet.max_column+1):
        d=sheet.cell(row=i,column=j)
        print(d.value)

#если зайти за пределы, то значение ячейки будет None
#Добавляем в файл в D2 значение, посмотрим, как будут себя вести значения ячеек
#передастся None там, где пусто, D2 будет корректно

#Есть и другой способ обхода книги
#диапазон
print('Диапазон')
for cellObj in sheet['A1':'D3']:
    for cell in cellObj:
        print(cell.value)
#cellObj - строка
#Остальное из openpyxl - гуглить
