#Работа с листами таблицы
#Книга из двух листов, на первом - ФИО и зарплата, на втором ФИО и регион
#надо посчитать среднюю зарплату по региону
from openpyxl import load_workbook
wb=load_workbook('test6-6-1.xlsx')
#sheet=wb.get_sheet_by_name('Лист1')
sheet=wb['Лист1']
salaries={}
#мы должны идти по всем строкам
for i in range(1,sheet.max_row+1):
    name=sheet.cell(row=i,column=1).value
    salary=sheet.cell(row=i,column=2).value
    salaries[name]=salary
print(salaries)
#смотрим на регионы (Лист2)
sheet=wb['Лист2']
regions={}
for i in range(1,sheet.max_row+1):
    name=sheet.cell(row=i,column=1).value
    region=sheet.cell(row=i,column=2).value
    if region not in regions:
        regions[region]=[0,0]
    regions[region][0]+=1
    regions[region][1]+=salaries[name]

for region in regions:
    print(region,regions[region][1]/regions[region][0])

#Т.о. с помощью словарей можно сопоставлять данные из разных таблиц