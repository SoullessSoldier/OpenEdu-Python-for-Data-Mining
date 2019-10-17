a='''
Васю назначили завхозом в туристической группе и он подошёл к подготовке ответственно, 
составив справочник продуктов с указанием калорийности на 100 грамм, 
а также содержание белков, жиров и углеводов на 100 грамм продукта. 
Ему не удалось найти всю информацию, поэтому некоторые ячейки остались незаполненными (можно считать их значение равным нулю). 
Также он использовал какой-то странный офисный пакет и разделял целую и дробную часть чисел запятой. 
Таблица доступна по ссылке https://stepik.org/media/attachments/lesson/245290/trekking2.xlsx 

Вася составил раскладку по продуктам на один день (она на листе "Раскладка") 
с указанием названия продукта и его количества в граммах. 
Посчитайте 4 числа: суммарную калорийность и граммы белков, жиров и углеводов. 
Числа округлите до целых вниз и введите через пробел.
'''
def replace_none(s):
    s=str(s)
    s.replace(',','.')
    if s == 'None':
        s=0
    s=float(s)

    return s
from math import floor

from openpyxl import load_workbook
wb=load_workbook('trekking6-6-2.xlsx')
sheet=wb['Раскладка']
raskladka1=[]
for i in range(2,sheet.max_row+1):
    raskladka1.append([sheet.cell(row=i,column=1).value,sheet.cell(row=i,column=2).value/100])
#print(products)
raskladka=dict()
sheet=wb['Справочник']
for i in range(2,sheet.max_row+1):
    raskladka[sheet.cell(row=i,column=1).value]=[replace_none(sheet.cell(row=i,column=2).value),
    replace_none(sheet.cell(row=i,column=3).value),
    replace_none(sheet.cell(row=i,column=4).value),
    replace_none(sheet.cell(row=i,column=5).value)]
#print(raskladka)
food_list=[]
for item in raskladka1:
    food_list.append([item[0],item[1]*raskladka[item[0]][0],item[1]*raskladka[item[0]][1],
                      item[1]*raskladka[item[0]][2],item[1]*raskladka[item[0]][3]])
for i in food_list:
    print(i)
kcal=0
b=0
z=0
u=0
for i in range(0,len(food_list)):
    kcal+=food_list[i][1]
    b+=food_list[i][2]
    z+=food_list[i][3]
    u+=food_list[i][4]
print(*map(floor,(kcal,b,z,u)),sep=' ')
#Извращенцы, тфу